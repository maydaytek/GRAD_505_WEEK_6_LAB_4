"""
GRAD 505 - Week 6 Lab 4: Risk Assessment in Food Science
=========================================================
Merges consumption, pesticide residue, and LD50 data to calculate
cumulative pesticide exposure risk by demographic group.

Outputs (written to ./outputs/):
  - Pesticide_Residue_Updated.xlsx   : original residue file + mean rows
  - Pesticide_Residue_Subset.xlsx    : trimmed to 5 target pesticides
  - Consumption_Revised.xlsx         : full merged/calculated dataset
  - Risk_Summary.csv                 : group-level mean +/- SD and rankings
"""

import os
import warnings
import numpy as np
import pandas as pd
from openpyxl import load_workbook

# Suppress only the pandas PerformanceWarning raised when adding many columns
warnings.filterwarnings("ignore", category=pd.errors.PerformanceWarning)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
MATERIALS = os.path.join(os.path.dirname(__file__), "MATERIALS")
OUTPUTS   = os.path.join(os.path.dirname(__file__), "outputs")
os.makedirs(OUTPUTS, exist_ok=True)

CONSUMPTION_FILE = os.path.join(MATERIALS, "Consumption1.xlsx")
PESTICIDE_FILE   = os.path.join(MATERIALS, "Pesticide Residue2.xlsx")
LD50_FILE        = os.path.join(MATERIALS, "LD50.xlsx")

# Reference dose for Disulfoton (U.S. EPA) in mg/kg/day
REFERENCE_DOSE = 0.0025
DAYS_PER_MONTH = 30.44

# ---------------------------------------------------------------------------
# Step 1 - Load LD50 table
# ---------------------------------------------------------------------------
print("Loading LD50 data...")
ld50_df = pd.read_excel(LD50_FILE, sheet_name="Sheet1")
ld50 = dict(zip(ld50_df["Pesticide"], ld50_df["LD50"]))
print(f"  LD50 values: {ld50}")

# ---------------------------------------------------------------------------
# Step 2 - Load pesticide residue data and calculate means
#   2015-Paper = Chinese kale  (data rows 3-119, 1-indexed)
#   2017-Paper = Cabbage       (data rows 4-85,  1-indexed)
#
# Raw units are ug/kg; divide by 1000 to convert to mg/kg.
# ---------------------------------------------------------------------------
print("\nLoading pesticide residue data...")

# Chinese kale (2015-Paper): header row is row 2 (0-indexed row 1); data rows 3-119
ck_raw = pd.read_excel(PESTICIDE_FILE, sheet_name="2015-Paper ",
                       header=1, nrows=117)
ck_raw.columns = [str(c).strip() for c in ck_raw.columns]
ck_raw = ck_raw.rename(columns={"Sample No.": "Sample_No"})

# Cabbage (2017-Paper): title in row 1, blank row 2, header row 3; data rows 4-85
cab_raw = pd.read_excel(PESTICIDE_FILE, sheet_name="2017-Paper",
                        header=2, nrows=82)
cab_raw.columns = [str(c).strip() for c in cab_raw.columns]
cab_raw = cab_raw.rename(columns={
    "Sample no.": "Sample_No",
    "λ-Cyhalothrin": "Cyhalothrin",
})

# Target pesticides per vegetable (must exist in LD50 table)
CK_PESTS  = ["Diazinon", "Dimethoate", "Malathion", "Profenofos"]
CAB_PESTS = ["Diazinon", "Dichlorvos", "Dimethoate"]


def mean_ug_to_mg(df, pesticides):
    """Return dict of mean concentrations in mg/kg (converted from ug/kg)."""
    means = {}
    for p in pesticides:
        col = df[p].dropna()
        means[p] = col.mean() / 1000.0  # ug/kg -> mg/kg
    return means


ck_means  = mean_ug_to_mg(ck_raw,  CK_PESTS)
cab_means = mean_ug_to_mg(cab_raw, CAB_PESTS)

print("  Chinese kale mean concentrations (mg/kg):")
for p, v in ck_means.items():
    print(f"    {p}: {v:.6f}")
print("  Cabbage mean concentrations (mg/kg):")
for p, v in cab_means.items():
    print(f"    {p}: {v:.6f}")

# ---------------------------------------------------------------------------
# Step 3 - Write Pesticide_Residue_Updated.xlsx (means appended)
# ---------------------------------------------------------------------------
print("\nWriting Pesticide_Residue_Updated.xlsx...")

wb_res = load_workbook(PESTICIDE_FILE)


def append_means_to_sheet(ws, pesticides, header_row, last_data_row):
    """Append mean rows (in original ug/kg and in mg/kg) below data."""
    col_map = {str(cell.value).strip(): cell.column
               for cell in ws[header_row] if cell.value}

    next_row = last_data_row + 2
    ws.cell(next_row, 1).value = "Mean (ug/kg)"
    for p in pesticides:
        if p in col_map:
            col = col_map[p]
            vals = [ws.cell(r, col).value
                    for r in range(header_row + 1, last_data_row + 1)
                    if ws.cell(r, col).value is not None]
            ws.cell(next_row, col).value = round(np.mean(vals), 4) if vals else None

    ws.cell(next_row + 1, 1).value = "Mean (mg/kg)"
    for p in pesticides:
        if p in col_map:
            col = col_map[p]
            src = ws.cell(next_row, col).value
            ws.cell(next_row + 1, col).value = (
                round(src / 1000.0, 6) if src is not None else None)


# 2015-Paper: header row 2 (1-indexed), data rows 3-119
ws_ck = wb_res["2015-Paper "]
append_means_to_sheet(ws_ck, CK_PESTS, header_row=2, last_data_row=119)

# 2017-Paper: header row 3 (1-indexed), data rows 4-85
ws_cab = wb_res["2017-Paper"]
append_means_to_sheet(ws_cab, CAB_PESTS, header_row=3, last_data_row=85)

updated_path = os.path.join(OUTPUTS, "Pesticide_Residue_Updated.xlsx")
wb_res.save(updated_path)
print(f"  Saved: {updated_path}")

# ---------------------------------------------------------------------------
# Step 4 - Write Pesticide_Residue_Subset.xlsx (only target pesticide cols)
# ---------------------------------------------------------------------------
print("\nWriting Pesticide_Residue_Subset.xlsx...")

ck_subset  = ck_raw[["Sample_No"] + CK_PESTS].copy()
cab_subset = cab_raw[["Sample_No"] + CAB_PESTS].copy()

# Append mean rows using pd.concat (preferred over df.loc[len(df)] in modern pandas)
for subset, means in [(ck_subset, ck_means), (cab_subset, cab_means)]:
    mean_row = pd.DataFrame([{"Sample_No": "Mean (mg/kg)", **means}])
    subset = pd.concat([subset, mean_row], ignore_index=True)

subset_path = os.path.join(OUTPUTS, "Pesticide_Residue_Subset.xlsx")
with pd.ExcelWriter(subset_path, engine="openpyxl") as writer:
    ck_subset.to_excel(writer,  sheet_name="2015-Paper_CK",  index=False)
    cab_subset.to_excel(writer, sheet_name="2017-Paper_Cab", index=False)
print(f"  Saved: {subset_path}")

# ---------------------------------------------------------------------------
# Step 5 - Load and prepare Consumption data
# ---------------------------------------------------------------------------
print("\nLoading consumption data...")
cons = pd.read_excel(CONSUMPTION_FILE, sheet_name="Sheet1")

# pandas auto-deduplicates repeated "Frequency" columns as Frequency, Frequency.1, ...
# Identify the correct frequency columns by position (column immediately after vegetable)
cols = list(cons.columns)
ck_freq_col  = cols[cols.index("Chinese kale") + 1]
cab_freq_col = cols[cols.index("Cabbage") + 1]

cons = cons.rename(columns={
    "Chinese kale": "CK_g",
    ck_freq_col:    "Freq_CK",
    "Cabbage":      "Cab_g",
    cab_freq_col:   "Freq_Cab",
    "Body weight":  "Body_weight",
})

# Keep only needed columns
keep = ["ID", "Sex", "Age", "Body_weight", "CK_g", "Freq_CK", "Cab_g", "Freq_Cab"]
df = cons[keep].copy()

print(f"  {len(df)} subjects loaded.")

# ---------------------------------------------------------------------------
# Step 6 - Build revised sheet calculations
# ---------------------------------------------------------------------------
print("\nBuilding revised dataset...")


def age_class(age):
    """Classify age into one of four groups: 1=<10, 2=10-17, 3=18-39, 4=40+."""
    if age < 10:
        return 1
    elif age < 18:
        return 2
    elif age < 40:
        return 3
    else:
        return 4


df["Age_class"]     = df["Age"].apply(age_class)
df["Sex_Age_class"] = (df["Sex"] - 1) * 4 + df["Age_class"]
# Sex=1 (Male): classes 1-4; Sex=2 (Female): classes 5-8

# Amount eaten per month (g/month = g/occasion * occasions/month)
df["Amount_Eaten_CK"]  = df["CK_g"]  * df["Freq_CK"]
df["Amount_Eaten_Cab"] = df["Cab_g"] * df["Freq_Cab"]

# Mean pesticide concentrations (broadcast the same scalar to every row)
for p in CK_PESTS:
    df[f"MC_{p}_CK"] = ck_means[p]
for p in CAB_PESTS:
    df[f"MC_{p}_Cab"] = cab_means[p]

# RPF = LD50 value for each pesticide (used as relative potency factor)
all_pests = sorted(set(CK_PESTS + CAB_PESTS))
for p in all_pests:
    df[f"RPF_{p}"] = ld50[p]

# TER = mean pesticide residue * LD50 (per pesticide-vegetable combination)
for p in CK_PESTS:
    df[f"TER_{p}_CK"]  = df[f"MC_{p}_CK"]  * df[f"RPF_{p}"]
for p in CAB_PESTS:
    df[f"TER_{p}_Cab"] = df[f"MC_{p}_Cab"] * df[f"RPF_{p}"]

# EXP = TER * Amount_Eaten / body_weight
for p in CK_PESTS:
    df[f"EXP_{p}_CK"]  = df[f"TER_{p}_CK"]  * df["Amount_Eaten_CK"]  / df["Body_weight"]
for p in CAB_PESTS:
    df[f"EXP_{p}_Cab"] = df[f"TER_{p}_Cab"] * df["Amount_Eaten_Cab"] / df["Body_weight"]

# SUM_EXP: total monthly exposure index per vegetable
df["SUM_EXP_CK"]    = df[[f"EXP_{p}_CK"  for p in CK_PESTS]].sum(axis=1)
df["SUM_EXP_Cab"]   = df[[f"EXP_{p}_Cab" for p in CAB_PESTS]].sum(axis=1)
df["SUM_EXP_Total"] = df["SUM_EXP_CK"] + df["SUM_EXP_Cab"]

# Relative exposure: convert monthly total to daily, then express as % of reference dose
ref_dose_monthly = REFERENCE_DOSE * DAYS_PER_MONTH
df["REL_EXP_CK_pct"]    = (df["SUM_EXP_CK"]   / ref_dose_monthly) * 100
df["REL_EXP_Cab_pct"]   = (df["SUM_EXP_Cab"]  / ref_dose_monthly) * 100
df["REL_EXP_Total_pct"] = (df["SUM_EXP_Total"] / ref_dose_monthly) * 100

print("  Calculations complete.")

# ---------------------------------------------------------------------------
# Step 7 - Save Consumption_Revised.xlsx
# ---------------------------------------------------------------------------
print("\nWriting Consumption_Revised.xlsx...")
revised_path = os.path.join(OUTPUTS, "Consumption_Revised.xlsx")
with pd.ExcelWriter(revised_path, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="revised", index=False)
print(f"  Saved: {revised_path}")

# ---------------------------------------------------------------------------
# Step 8 - Risk Summary by group
# ---------------------------------------------------------------------------
print("\nGenerating Risk_Summary.csv...")

group_labels = {
    1: "Male <10",
    2: "Male 10-17",
    3: "Male 18-39",
    4: "Male 40+",
    5: "Female <10",
    6: "Female 10-17",
    7: "Female 18-39",
    8: "Female 40+",
}

summary_rows = []
for g in range(1, 9):
    grp = df[df["Sex_Age_class"] == g]["REL_EXP_Total_pct"]
    summary_rows.append({
        "Group_ID":       g,
        "Group_Label":    group_labels[g],
        "N":              len(grp),
        "Mean_REL_EXP_%": round(grp.mean(), 4),
        "SD_REL_EXP_%":   round(grp.std(),  4),
        "Min_%":          round(grp.min(),  4),
        "Max_%":          round(grp.max(),  4),
    })

summary_df = pd.DataFrame(summary_rows).sort_values("Mean_REL_EXP_%", ascending=False)
summary_df["Rank"] = range(1, len(summary_df) + 1)

overall_mean = df["REL_EXP_Total_pct"].mean()
overall_sd   = df["REL_EXP_Total_pct"].std()

print(f"\n  Overall mean relative exposure: {overall_mean:.4f}%")
print(f"  Overall SD:                     {overall_sd:.4f}%")
print("\n  Risk ranking by group:")
for _, row in summary_df.iterrows():
    print(f"    #{int(row['Rank'])}: {row['Group_Label']:15s}  "
          f"mean={row['Mean_REL_EXP_%']:.4f}%  SD={row['SD_REL_EXP_%']:.4f}%  "
          f"n={int(row['N'])}")

summary_path = os.path.join(OUTPUTS, "Risk_Summary.csv")
summary_df.to_csv(summary_path, index=False)
print(f"\n  Saved: {summary_path}")

# ---------------------------------------------------------------------------
# Done
# ---------------------------------------------------------------------------
print("\nAll outputs written to ./outputs/")
print("  - Pesticide_Residue_Updated.xlsx")
print("  - Pesticide_Residue_Subset.xlsx")
print("  - Consumption_Revised.xlsx")
print("  - Risk_Summary.csv")
